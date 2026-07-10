"""Probe public Osaka / Kansai grammar data sources for koyori lexicon work.

Reads archives under .research/osaka-grammar-data/ (see docs/tracks/osaka-grammar-data.md).
Writes UTF-8 JSON summaries to .research/osaka-grammar-data/out/.
"""

from __future__ import annotations

import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / ".research" / "osaka-grammar-data"
OUT = DATA / "out"

# GAJ vol.2–3 workbook holds survey ticket 1 (negation 001–015, copula 028, etc.)
GAJ_VOL23 = DATA / "gaj_all" / "GAJ23_all_unicode+geocode_202306.xlsx"

GAJ_GRAMMAR_SETS: dict[str, set[str]] = {
    "negation": {f"{i:03d}" for i in range(1, 16)},
    "copula": {"028"},
    "existence": {"031", "032"},  # ある / ない (existential)
}

OSAKA_PREF_NAMES = ("大阪府",)

KVJ_CORE = DATA / "kvj" / "kvjcorpus-suw-0.9" / "kvjcorpus-suw-core.txt"
ITA_KANSAI = DATA / "ITA_KANSAI_ONRY.txt"

# Surface patterns for corpus frequency (Osaka-oriented young speech in KVJ/ITA).
SURFACE_PATTERNS: dict[str, re.Pattern[str]] = {
    "neg_hen": re.compile(r"へん$|ひん$"),
    "neg_sen": re.compile(r"せん$"),
    "neg_nai": re.compile(r"ない$|ねえ$"),
    "copula_ya": re.compile(r"^や$|やな$|やで$|やねん$|やろ$"),
    "sfp_wa": re.compile(r"わ$|わな$"),
    "sfp_de": re.compile(r"で$|でぇ$"),
    "sfp_ne": re.compile(r"ねん$|ん$"),
    "chau": re.compile(r"ちゃう|ちゃ|ちゅう"),
    "akan": re.compile(r"あかん"),
    "omounai": re.compile(r"おもんない|おもろない"),
    "dekihin": re.compile(r"でけへん|できひん|できへん"),
}


def _write(name: str, payload: object) -> Path:
    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / name
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _load_openpyxl():
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError:
        import subprocess
        import sys

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl  # type: ignore[import-not-found]
    return openpyxl


def gaj_osaka_point_codes() -> set[str]:
    openpyxl = _load_openpyxl()
    xlsx = DATA / "gaj_points" / "GAJ_ALL_PointProperty.xlsx"
    if not xlsx.is_file():
        return set()
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb.active
    codes: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        pref = str(row[3] or "")
        if any(name in pref for name in OSAKA_PREF_NAMES):
            codes.add(str(row[0]))
    wb.close()
    return codes


def gaj_research_items() -> dict[str, dict[str, str]]:
    path = DATA / "gaj_items" / "research_item.txt"
    if not path.is_file():
        return {}
    items: dict[str, dict[str, str]] = {}
    raw = path.read_bytes()
    for enc in ("utf-8", "cp932", "utf-8-sig"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            text = raw.decode("utf-8", errors="replace")
    lines = text.splitlines()
    for line in lines[1:]:
        cols = line.split("\t")
        if len(cols) < 5:
            continue
        qid = cols[1]
        items[qid] = {
            "category": cols[2],
            "label": cols[3],
            "question": cols[4],
            "map": cols[6] if len(cols) > 6 else "",
        }
    return items


def gaj_osaka_answers(
    *,
    qids: set[str],
    xlsx: Path = GAJ_VOL23,
    limit_per_q: int = 12,
) -> dict[str, object]:
    openpyxl = _load_openpyxl()
    if not xlsx.is_file():
        return {"error": f"missing {xlsx}"}
    osaka_codes = gaj_osaka_point_codes()
    items = gaj_research_items()
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(h): i for i, h in enumerate(header) if h}

    by_q: dict[str, Counter[str]] = defaultdict(Counter)
    for row in ws.iter_rows(min_row=2, values_only=True):
        qid = str(row[idx.get("質問番号", 1)] or "")
        if qid not in qids:
            continue
        point = str(row[idx.get("地点番号", 2)] or "")
        if point not in osaka_codes:
            continue
        form = str(row[idx.get("回答語形", 5)] or "").strip()
        if form and form != "無回答":
            by_q[qid][form] += 1

    wb.close()
    out: dict[str, object] = {
        "source_workbook": xlsx.name,
        "osaka_point_count": len(osaka_codes),
        "osaka_points": sorted(osaka_codes),
        "questions": {},
    }
    for qid, counter in sorted(by_q.items()):
        meta = items.get(qid, {})
        out["questions"][qid] = {
            **meta,
            "top_answers": counter.most_common(limit_per_q),
            "distinct_answers": len(counter),
        }
    return out


def gaj_grammar_bundle() -> dict[str, object]:
    bundle: dict[str, object] = {}
    for name, qids in GAJ_GRAMMAR_SETS.items():
        bundle[name] = gaj_osaka_answers(qids=qids)
    return bundle


def unidic_via_mecab(patterns: tuple[str, ...], limit: int = 120) -> dict[str, object]:
    """Extract morpheme entries using MeCab + local unidic_kansai dict."""
    dic_dir = DATA / "unidic_kansai"
    if not (dic_dir / "dicrc").is_file():
        return {"error": "unidic_kansai not extracted", "hits": []}

    try:
        import MeCab  # type: ignore[import-not-found]
    except ImportError:
        return {
            "error": "MeCab Python binding not installed (pip install mecab-python3 + MeCab binary)",
            "dic_dir": str(dic_dir),
            "hits": [],
            "workaround": "Use Web 茶まめ https://chamame.ninjal.ac.jp/ or install MeCab for Windows",
        }

    tagger = MeCab.Tagger(f"-d {dic_dir}")
    # Seed lookup: parse example sentences containing target morphemes.
    seeds = [
        "おもろないわ",
        "できへんねん",
        "あかんわ",
        "ほんまやな",
        "知らへん",
        "せえへん",
        "ちゃうちゃう",
        "おれやで",
    ]
    hits: list[dict[str, str]] = []
    seen: set[str] = set()
    for sent in seeds:
        node = tagger.parseToNode(sent)
        while node:
            surf = node.surface
            feat = node.feature
            if surf and any(p in surf for p in patterns) and surf not in seen:
                seen.add(surf)
                hits.append({"surface": surf, "feature": feat, "seed": sent})
            node = node.next
            if len(hits) >= limit:
                break
    return {"hits": hits, "seed_count": len(seeds)}


def strip_adams_html_ruby(html_path: Path) -> str:
    if not html_path.is_file():
        return ""
    html = html_path.read_text(encoding="utf-8", errors="replace")
    html = re.sub(r"<rt[^>]*>.*?</rt>", "", html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r"<rp[^>]*>.*?</rp>", "", html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r"<[^>]+>", "\n", html)
    html = re.sub(r"\n{3,}", "\n\n", html)
    lines = [ln.strip() for ln in html.splitlines() if ln.strip()]
    return "\n".join(lines)


def adams_grammar_index(plain: str) -> list[dict[str, object]]:
    """Chapter titles + example lines with Osaka-relevant grammar glosses."""
    chapter_re = re.compile(
        r"^(Negative Verb Conjugation|Plain Copula|Sentence Final Particle|"
        r"Expression|\"Wrong;\"|\"Bad;\"|\"To Say;\"|\"Good;\"|\"Interesting;\"|"
        r"To Be, To Exist|Genderized Sentence Final|Contracted\s+Adjectives|"
        r"vs\.\s+.*Insults)\s*$",
        flags=re.IGNORECASE | re.MULTILINE,
    )
    chapters: list[dict[str, object]] = []
    for m in chapter_re.finditer(plain):
        chapters.append({"title": m.group(1).strip(), "offset": m.start()})

    example_re = re.compile(
        r"(できへん|できひん|あかん|ちゃう|ほんま|おもろない|おもんない|"
        r"知らへん|せえへん|やな|やで|やねん|やろ|ちゃうねん|うてへん)",
    )
    examples: Counter[str] = Counter()
    for m in example_re.finditer(plain):
        examples[m.group(1)] += 1

    return [{"chapters": chapters}, {"example_tokens": examples.most_common(40)}]


def mine_kvj_ksj(*, subcorpus_prefix: str = "KSJ") -> dict[str, object]:
    if not KVJ_CORE.is_file():
        return {"error": f"missing {KVJ_CORE}"}

    pattern_counts: dict[str, Counter[str]] = {k: Counter() for k in SURFACE_PATTERNS}
    file_ids: set[str] = set()
    token_total = 0

    with KVJ_CORE.open(encoding="utf-8", errors="replace") as fh:
        for line in fh:
            cols = line.rstrip("\n").split("\t")
            if len(cols) < 6:
                continue
            file_id = cols[0]
            if not file_id.startswith(subcorpus_prefix):
                continue
            file_ids.add(file_id)
            surface = cols[5]
            token_total += 1
            for key, pat in SURFACE_PATTERNS.items():
                if pat.search(surface):
                    pattern_counts[key][surface] += 1

    return {
        "subcorpus": subcorpus_prefix,
        "file_count": len(file_ids),
        "token_count": token_total,
        "pattern_hits": {
            k: c.most_common(15) for k, c in sorted(pattern_counts.items()) if c
        },
        "note": "KSJ = Osaka/Kobe urban; may include some Kyoto/Hyogo speakers per KVJ docs",
    }


def mine_ita_kansai() -> dict[str, object]:
    if not ITA_KANSAI.is_file():
        return {"error": f"missing {ITA_KANSAI}"}
    lines = [
        ln.strip()
        for ln in ITA_KANSAI.read_text(encoding="utf-8", errors="replace").splitlines()
        if ln.strip()
    ]
    pattern_lines: dict[str, list[str]] = defaultdict(list)
    for ln in lines:
        for key, pat in SURFACE_PATTERNS.items():
            if pat.search(ln) and len(pattern_lines[key]) < 8:
                pattern_lines[key].append(ln)
    return {
        "source": "ITA_KANSAI_ONRY.txt (Osaka-flavored TTS sentences, GitHub)",
        "line_count": len(lines),
        "sample_by_pattern": dict(pattern_lines),
    }


def build_grammar_sketch(
    gaj: dict[str, object],
    kvj: dict[str, object],
) -> dict[str, object]:
    """Merge top evidence into a compose-friendly sketch (not prescriptive rules)."""
    sketch: dict[str, object] = {"rules": []}

    neg = gaj.get("negation", {})
    if isinstance(neg, dict):
        q = neg.get("questions", {})
        if isinstance(q, dict) and "004" in q:
            item = q["004"]
            if isinstance(item, dict):
                tops = item.get("top_answers", [])
                sketch["rules"].append(
                    {
                        "standard": "しない",
                        "category": "negation",
                        "gaj_osaka_forms": tops,
                        "gaj_label": item.get("label"),
                        "confidence": "low (n=6 GAJ points)",
                    }
                )

    kvj_hits = kvj.get("pattern_hits", {})
    if isinstance(kvj_hits, dict):
        for key in ("neg_hen", "dekihin", "omounai", "akan", "copula_ya"):
            hits = kvj_hits.get(key)
            if hits:
                sketch["rules"].append(
                    {
                        "pattern_key": key,
                        "kvj_ksj_top_surfaces": hits[:8],
                        "source": "KVJ core KSJ",
                    }
                )
    return sketch


def main() -> int:
    summary: dict[str, object] = {"data_dir": str(DATA), "outputs": []}

    gaj = gaj_grammar_bundle()
    p_gaj = _write("gaj_osaka_grammar.json", gaj)
    summary["outputs"].append(str(p_gaj))

    kvj = mine_kvj_ksj()
    p_kvj = _write("kvj_ksj_patterns.json", kvj)
    summary["outputs"].append(str(p_kvj))

    ita = mine_ita_kansai()
    p_ita = _write("ita_kansai_samples.json", ita)
    summary["outputs"].append(str(p_ita))

    morphemes = {
        "negation_related": unidic_via_mecab(("へん", "ひん", "せん", "あかん")),
        "note": "Full UniDic lexicon requires MeCab; sys.dic is binary",
    }
    p_uni = _write("unidic_kansai_morphemes.json", morphemes)
    summary["outputs"].append(str(p_uni))

    html = ROOT / "docs" / "KansaibenSurvivalManual_AdamsZach2010_.html"
    plain = strip_adams_html_ruby(html)
    if plain:
        p_plain = OUT / "adams_plain.txt"
        p_plain.write_text(plain[:200_000], encoding="utf-8")
        adams = adams_grammar_index(plain)
        p_adams = _write("adams_grammar_index.json", adams)
        summary["outputs"].append(str(p_plain))
        summary["outputs"].append(str(p_adams))

    sketch = build_grammar_sketch(gaj, kvj)
    p_sketch = _write("osaka_grammar_sketch.json", sketch)
    summary["outputs"].append(str(p_sketch))

    notes = {
        "kvj_corpus": {
            "short_unit": {
                "url": "https://repository.ninjal.ac.jp/record/2000486/files/kvjcorpus-suw-0.9.zip",
                "license": "CC BY-NC-SA 4.0",
                "osaka_subcorpus": "KSJ* file ids in kvjcorpus-suw-core.txt",
            },
            "raw_interviews": {
                "url": "https://sites.google.com/view/kvjcorpus",
                "osaka_zip": "KSJ.zip — manual agree on site",
                "license": "CC BY-NC-SA 4.0",
            },
        },
        "unidic_kansai": {
            "url": "https://clrd.ninjal.ac.jp/unidic/download_all.html",
            "file": "unidic-D-kansai-v202512.zip",
            "license": "CC BY-NC-SA 4.0",
            "usage": "MeCab -d unidic_kansai/ ; or Web 茶まめ",
        },
        "gaj": {
            "url": "https://www2.ninjal.ac.jp/gaj/",
            "osaka_filter": "GAJ_ALL_PointProperty.xlsx 都道府県=大阪府 (6 points)",
            "negation_workbook": "GAJ23_all_unicode+geocode_202306.xlsx",
            "license": "research use; cite 方言文法全国地図",
        },
        "cojads": {
            "url": "https://www2.ninjal.ac.jp/cojads/",
            "free": "メタ情報 CSV — COJADS site データ配布ページ; 中納言登録要",
            "parallel": "標準語 + 方言カタカナ",
            "osaka": "大阪地点あり（1977–85 調査・口語は古い）",
            "caveat": "LoRA/compose には年代・演技談話に注意",
        },
        "adams_manual": {
            "local": "docs/KansaibenSurvivalManual_AdamsZach2010_.html",
            "site": "https://kansaiben.com/",
            "caveat": "ルビ付きHTMLは機械処理向きでない → adams_plain.txt",
        },
    }
    p_notes = _write("source_notes.json", notes)
    summary["outputs"].append(str(p_notes))

    p0 = _write("run_summary.json", summary)
    print(p0)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
